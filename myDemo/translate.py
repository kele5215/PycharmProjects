import re
import urllib.parse, urllib.request
import hashlib
import urllib
import random
import json
import time
from googletrans import Translator
from openpyxl import load_workbook


def translate_google_2(text_str):
    translator = Translator()
    detect_result = translator.detect(text_str)
    if detect_result.lang == 'en':
        return detect_result.text
    result = translator.translate(text_str)
    result = translator.translate(result.text, dest='zh-CN')
    # print(result.text)
    return result.text


def do_excel(path_excel):
    wb = load_workbook(path_excel)
    # print(wb.sheetnames)
    sheet = wb["ResXResourceManager"]

    # sheet_row = sheet.max_row
    # sheet_column = sheet.max_column

    # print("sheet_row:", sheet_row)
    # print("sheet_column:", sheet_column)

    for row_item in sheet["E"]:
        if row_item.row == 1:
            continue
        if not row_item.value or len(row_item.value) == 0:
            continue
        # print(row_item.value)
        translated_value = translate_google_2(row_item.value)

        sheet["G%d" % row_item.row].value = translated_value
        time.sleep(3)
    wb.save(path_excel)


# def translate_google(text, f='ja', t='zh-cn'):
#     url_google = 'https://translate.google.com'
#     reg_text = re.compile(r'(?<=TRANSLATED_TEXT=).*?;')
#     user_agent = r'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
#                  r'Chrome/44.0.2403.157 Safari/537.36'
#
#     values = {'client': 'webapp',
#               'sl': f,
#               'tl': t,
#               'hl': 'ja',
#               'dt': 'at',
#               'dt': 'bd',
#               'dt': 'ex',
#               'dt': 'ld',
#               'dt': 'md',
#               'dt': 'qca',
#               'dt': 'rw',
#               'dt': 'rm',
#               'dt': 'ss',
#               'dt': 't',
#               'source': 'bh',
#               'ssel': '0',
#               'tsel': '0',
#               'kc': '1',
#               'tk': '766023.862052',
#               'q': text, }
#     value = urllib.parse.urlencode(values)
#     req = urllib.request.Request(url_google + '/translate_a/single?' + value)
#     req.add_header('User-Agent', user_agent)
#     response = urllib.request.urlopen(req)
#     content = response.read().decode('utf-8')
#     target = json.loads(content)
#     tgt = target['trans_result']['data'][0]['dst']
#     data = reg_text.search(content)
#     result = data.group(0).strip(';').strip('\'')
#     print(result)


if __name__ == '__main__':
    time1 = time.time()
    do_excel(r'C:\Users\h-zhang\Desktop\temp\2019_04_11_16_22.xlsx')
    time2 = time.time()

    print('Googletransで通訳時間：%s' % (time2 - time1))
